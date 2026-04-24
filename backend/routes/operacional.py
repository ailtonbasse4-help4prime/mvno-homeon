"""Modulo Operacional - Visao consolidada tipo planilha (Excel-like)."""
import io
import re
import unicodedata
import logging
from datetime import datetime, timezone, timedelta
from typing import Optional, List
from fastapi import APIRouter, Request, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from bson import ObjectId
from pydantic import BaseModel

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/operacional", tags=["operacional"])

# Injetado pelo server.py principal
_ctx = {}


def init(db, get_current_user, require_admin, create_log):
    _ctx["db"] = db
    _ctx["get_current_user"] = get_current_user
    _ctx["require_admin"] = require_admin
    _ctx["create_log"] = create_log


def _parse_data_br(s):
    """Converte DD-MM-YYYY ou DD/MM/YYYY em YYYY-MM-DD. Aceita ISO tb."""
    if not s:
        return None
    s = str(s).strip()[:10]
    import re as _re
    m = _re.match(r'^(\d{2})[-/](\d{2})[-/](\d{4})$', s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    if _re.match(r'^\d{4}-\d{2}-\d{2}$', s):
        return s
    return s


def _calc_proxima_recarga(data_ativacao_iso, data_bloqueio_iso):
    """Calcula a data da PROXIMA RECARGA do chip na Ta Telecom.

    A API da Ta Telecom retorna `data_ativacao` e `data_bloqueio`. Quando o
    chip nunca foi bloqueado, `data_bloqueio` vem igual a `data_ativacao` ou null.
    Quando o chip vai ser bloqueado (recarga programada), `data_bloqueio` traz
    uma data futura (a proxima recarga). Logica:
      - Se data_bloqueio > hoje: usar data_bloqueio (recarga programada pela Ta).
      - Caso contrario: calcular a proxima mensalidade a partir da data de
        ativacao, avancando 30 dias por ciclo ate encontrar uma data futura.
        Assim funciona para chips ativos ha varios meses sem data_bloqueio.
    """
    from datetime import datetime as _dt, date as _date, timedelta as _td
    def _pd(s):
        if not s: return None
        try:
            return _dt.strptime(s[:10], "%Y-%m-%d").date()
        except Exception:
            return None
    hoje = _date.today()
    at = _pd(data_ativacao_iso)
    bl = _pd(data_bloqueio_iso)
    # 1. Se data_bloqueio da Ta e FUTURA, ela e a proxima recarga real
    if bl and bl > hoje:
        return bl.strftime("%Y-%m-%d")
    # 2. Calcular ciclo mensal a partir da ativacao
    if at:
        prox = at + _td(days=30)
        # Avanca 30 dias ate encontrar data futura (chip ativo ha varios meses)
        while prox <= hoje:
            prox = prox + _td(days=30)
        return prox.strftime("%Y-%m-%d")
    # 3. Fallback: data_bloqueio mesmo que passada (ultima info disponivel)
    if bl:
        return bl.strftime("%Y-%m-%d")
    return None


def _resolver_proxima_recarga(*, data_ativacao, data_bloqueio, data_expiracao_direta):
    """Resolve a PROXIMA RECARGA considerando todos os candidatos da API Ta.

    Regra mestre: a proxima recarga SEMPRE tem que ser uma data FUTURA (>= hoje).
    Se a API retornar qualquer data no passado (ex: data_expiracao do ciclo
    anterior), descarta e calcula via ciclo de 30d a partir da ativacao.
    """
    from datetime import datetime as _dt, date as _date, timedelta as _td
    def _pd(s):
        if not s: return None
        try:
            return _dt.strptime(s[:10], "%Y-%m-%d").date()
        except Exception:
            return None
    hoje = _date.today()
    at = _pd(data_ativacao)
    bl = _pd(data_bloqueio)
    ex = _pd(data_expiracao_direta)

    # 1. data_expiracao explicita, se for FUTURA - melhor candidata
    if ex and ex >= hoje:
        return ex.strftime("%Y-%m-%d")
    # 2. data_bloqueio futura - proxima recarga programada
    if bl and bl > hoje:
        return bl.strftime("%Y-%m-%d")
    # 3. Calcular ciclo mensal a partir da ativacao (avanca 30d ate achar data futura)
    if at:
        prox = at + _td(days=30)
        while prox <= hoje:
            prox = prox + _td(days=30)
        return prox.strftime("%Y-%m-%d")
    # 4. Fallback: data_expiracao passada, depois data_bloqueio passada
    if ex:
        return ex.strftime("%Y-%m-%d")
    if bl:
        return bl.strftime("%Y-%m-%d")
    return None


def _parse_tamanho_gb(nome: str, franquia: str = "") -> float:
    """Extrai tamanho em GB do nome do plano/franquia. MB -> GB fracionado."""
    text = f"{nome} {franquia}".upper()
    m = re.search(r'(\d+(?:[\.,]\d+)?)\s*(GB|MB)', text)
    if not m:
        return 999999  # sem tamanho detectado vai pro final
    n = float(m.group(1).replace(',', '.'))
    return n if m.group(2) == 'GB' else n / 1024


def _sort_key_plano(nome: str, franquia: str = ""):
    """Ordena planos: M2M por ultimo, resto por GB crescente."""
    text = f"{nome} {franquia}".upper()
    is_m2m = 1 if 'M2M' in text else 0
    return (is_m2m, _parse_tamanho_gb(nome, franquia))


def _norm(s: str) -> str:
    if not s:
        return ""
    return ''.join(c for c in unicodedata.normalize('NFD', str(s)) if unicodedata.category(c) != 'Mn').lower()


class LinhaOperacionalUpdate(BaseModel):
    observacoes: Optional[str] = None
    proxima_recarga: Optional[str] = None  # ISO date YYYY-MM-DD
    canal: Optional[str] = None  # atualizar no cliente
    status_chip: Optional[str] = None  # texto livre
    complemento: Optional[str] = None  # identificacao interna (ex: "filho Joao")
    incluir_custo: Optional[bool] = None  # se deve contar custo no total
    incluir_lucro: Optional[bool] = None  # se deve contar receita no lucro
    desconto: Optional[float] = None  # desconto fixo em R$ sobre o valor do plano (ex: combo)


class OfertaCustoUpdate(BaseModel):
    custo: float


class CustosBatchUpdate(BaseModel):
    custos: dict  # {"oferta_id": custo_float}


class PlanoCustoUpdate(BaseModel):
    custo: float  # aplica a TODAS as ofertas desse plano


class CustoFixoCreate(BaseModel):
    nome: str
    valor: float
    ativo: bool = True


class CustoFixoUpdate(BaseModel):
    nome: Optional[str] = None
    valor: Optional[float] = None
    ativo: Optional[bool] = None


@router.get("/planilha")
async def planilha_consolidada(request: Request, search: Optional[str] = None, status: Optional[str] = None,
                                canal: Optional[str] = None, bloqueio: Optional[str] = None):
    """Retorna uma linha por LINHA ativa/suspensa com dados consolidados: cliente+chip+oferta+plano+cobrancas."""
    await _ctx["get_current_user"](request)
    db = _ctx["db"]

    # Auto-sync com Ta Telecom em BACKGROUND para linhas com data passada ou stale.
    # Dispara no maximo 1x a cada 10 minutos (usa flag no mongodb como debounce).
    try:
        import asyncio as _asyncio
        from datetime import date as _date, timedelta as _td
        hoje_iso_str = _date.today().strftime("%Y-%m-%d")
        # Debounce: verifica se ja sincronizou nos ultimos 10 min
        flag = await db.sync_flags.find_one({"key": "planilha_auto_sync"})
        now = datetime.now(timezone.utc)
        cooldown_ok = True
        if flag and flag.get("last_run"):
            last = flag["last_run"]
            if isinstance(last, datetime):
                if (now - last).total_seconds() < 600:
                    cooldown_ok = False
        if cooldown_ok:
            # Pega linhas com data passada OU sem data OU nao-sincronizadas ha >12h
            cutoff = now - timedelta(hours=12)
            stale = await db.linhas.find({
                "$or": [
                    {"expirar_dados": None},
                    {"expirar_dados": {"$lt": hoje_iso_str}},
                    {"expirar_dados_updated_at": {"$lt": cutoff}},
                    {"expirar_dados_updated_at": {"$exists": False}},
                ],
            }, {"_id": 1, "chip_id": 1}).limit(200).to_list(200)
            stale_ids = [str(l["_id"]) for l in stale]
            if stale_ids:
                await db.sync_flags.update_one(
                    {"key": "planilha_auto_sync"},
                    {"$set": {"last_run": now, "count": len(stale_ids)}},
                    upsert=True,
                )
                # Dispara sync em background (nao bloqueia a resposta)
                _asyncio.create_task(_auto_sync_linhas_stale(stale_ids))
    except Exception as _e:
        logger.warning(f"Falha no auto-sync da planilha (nao-critico): {_e}")

    # Buscar todas as linhas
    linhas = await db.linhas.find({}).to_list(5000)
    if not linhas:
        return {"linhas": [], "resumo": _resumo_vazio()}

    cliente_ids = list({l["cliente_id"] for l in linhas if l.get("cliente_id")})
    chip_ids = list({l["chip_id"] for l in linhas if l.get("chip_id") and ObjectId.is_valid(l["chip_id"])})
    oferta_ids = list({l["oferta_id"] for l in linhas if l.get("oferta_id") and ObjectId.is_valid(l["oferta_id"])})
    plano_ids_linha = list({l["plano_id"] for l in linhas if l.get("plano_id") and ObjectId.is_valid(l["plano_id"])})

    clientes_map = {}
    if cliente_ids:
        cs = await db.clientes.find({"_id": {"$in": [ObjectId(c) for c in cliente_ids if ObjectId.is_valid(c)]}}).to_list(5000)
        clientes_map = {str(c["_id"]): c for c in cs}

    chips_map = {}
    if chip_ids:
        chs = await db.chips.find({"_id": {"$in": [ObjectId(c) for c in chip_ids]}}).to_list(5000)
        chips_map = {str(c["_id"]): c for c in chs}

    ofertas_map = {}
    planos_map = {}
    # Carrega todas ofertas ativas (para fallback por plano)
    all_ofertas = await db.ofertas.find({"ativo": True}).to_list(1000)
    for o in all_ofertas:
        ofertas_map[str(o["_id"])] = o
    # Indice: plano_id -> primeira oferta ativa
    oferta_por_plano = {}
    for o in all_ofertas:
        pid = o.get("plano_id")
        if pid and pid not in oferta_por_plano:
            oferta_por_plano[pid] = o

    all_plano_ids = list(set(plano_ids_linha) | {o.get("plano_id") for o in all_ofertas if o.get("plano_id")})
    if all_plano_ids:
        pls = await db.planos.find({"_id": {"$in": [ObjectId(p) for p in all_plano_ids if ObjectId.is_valid(p)]}}).to_list(500)
        planos_map = {str(p["_id"]): p for p in pls}

    # Ordenar cobrancas ASC (mais antigas primeiro) para facilitar selecao da "vencida mais antiga" e "proxima a vencer"
    cobs = await db.cobrancas.find({}).sort("vencimento", 1).to_list(10000)
    cobs_by_cliente = {}
    for c in cobs:
        cid = c.get("cliente_id")
        if not cid:
            continue
        if cid not in cobs_by_cliente:
            cobs_by_cliente[cid] = []
        cobs_by_cliente[cid].append(c)

    hoje_iso = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    PAID_STATUSES = ("RECEIVED", "CONFIRMED", "RECEIVED_IN_CASH", "REFUNDED")

    result = []
    for l in linhas:
        cid = l.get("cliente_id") or ""
        cliente = clientes_map.get(cid, {})
        chip = chips_map.get(l.get("chip_id") or "", {})
        oferta = ofertas_map.get(l.get("oferta_id") or "", {})
        # Fallback: se linha nao tem oferta, usa a primeira oferta ativa do plano da linha
        if not oferta and l.get("plano_id"):
            oferta = oferta_por_plano.get(l["plano_id"], {})
        plano = planos_map.get(oferta.get("plano_id") or l.get("plano_id") or "", {})
        cobs_cli = cobs_by_cliente.get(cid, [])
        # Cobrancas nao pagas ordenadas ASC por vencimento
        cobs_pend = [c for c in cobs_cli if c.get("status") not in PAID_STATUSES]
        # 1. Vencidas (vencimento < hoje) - mostra a MAIS ANTIGA (maior atraso)
        vencidas = [c for c in cobs_pend if (c.get("vencimento") or "") < hoje_iso]
        # 2. Futuras (vencimento >= hoje) - proxima a vencer
        futuras = [c for c in cobs_pend if (c.get("vencimento") or "") >= hoje_iso]
        if vencidas:
            ultima_cob = vencidas[0]  # mais antiga vencida (maior atraso)
        elif futuras:
            ultima_cob = futuras[0]  # proxima a vencer
        else:
            ultima_cob = cobs_cli[-1] if cobs_cli else {}  # sem pendentes: ultima paga
        total_cobs = len(cobs_cli)
        total_pagas = sum(1 for c in cobs_cli if c.get("status") in ("RECEIVED", "CONFIRMED", "RECEIVED_IN_CASH"))
        total_pendentes = total_cobs - total_pagas

        endereco_full = " ".join(filter(None, [
            cliente.get("endereco"), cliente.get("numero_endereco"),
            cliente.get("bairro"), cliente.get("cidade"), cliente.get("estado"),
        ]))

        valor = oferta.get("valor", 0.0) or 0.0
        custo = oferta.get("custo", 0.0) or 0.0
        desconto = float(l.get("desconto") or 0.0)
        if desconto < 0:
            desconto = 0.0
        if desconto > valor:
            desconto = valor
        valor_liquido = valor - desconto
        lucro = valor_liquido - custo
        margem = (lucro / valor_liquido * 100) if valor_liquido > 0 else 0

        # Flags de inclusao em custo e lucro (default: ativo=True, resto=False)
        status_linha = l.get("status", "")
        default_incluir = status_linha == "ativo"
        incluir_custo = l.get("incluir_custo")
        incluir_lucro = l.get("incluir_lucro")
        if incluir_custo is None:
            incluir_custo = default_incluir
        if incluir_lucro is None:
            incluir_lucro = default_incluir

        row = {
            "linha_id": str(l["_id"]),
            "cliente_id": cid,
            "chip_id": l.get("chip_id") or "",
            "oferta_id": l.get("oferta_id") or "",
            # Chip/Linha
            "iccid": chip.get("iccid", ""),
            "numero": l.get("numero") or l.get("msisdn") or chip.get("msisdn") or "",
            "status_linha": status_linha,
            "status_chip": l.get("status_chip") or chip.get("status", ""),
            "expirar_dados": l.get("expirar_dados"),
            "proxima_recarga": l.get("proxima_recarga"),
            "complemento": l.get("complemento", ""),
            "incluir_custo": incluir_custo,
            "incluir_lucro": incluir_lucro,
            # Cliente
            "cliente_nome": cliente.get("nome", ""),
            "cpf": cliente.get("documento") or cliente.get("cpf", ""),
            "telefone_cliente": cliente.get("telefone", ""),
            "email": cliente.get("email"),
            "endereco": endereco_full,
            "canal": cliente.get("canal", ""),
            "observacoes_cliente": cliente.get("observacoes", ""),
            # Oferta / Plano
            "oferta_nome": oferta.get("nome", ""),
            "plano_nome": plano.get("nome", ""),
            "franquia": plano.get("franquia", ""),
            "valor": valor,
            "desconto": round(desconto, 2),
            "valor_liquido": round(valor_liquido, 2),
            "custo": custo,
            "lucro": round(lucro, 2),
            "margem_pct": round(margem, 2),
            "categoria": oferta.get("categoria", ""),
            # Cobranca
            "ultima_cobranca_venc": ultima_cob.get("vencimento") if ultima_cob else None,
            "ultima_cobranca_status": ultima_cob.get("status") if ultima_cob else None,
            "ultima_cobranca_tipo": ultima_cob.get("billing_type") if ultima_cob else None,
            "cobrancas_total": total_cobs,
            "cobrancas_pagas": total_pagas,
            "cobrancas_pendentes": total_pendentes,
            # Observacoes da linha
            "observacoes_linha": l.get("observacoes", ""),
            "port": bool(cliente.get("port") or cliente.get("portabilidade")),
        }
        result.append(row)

    # Filtros
    if search:
        s = _norm(search)
        result = [r for r in result if s in _norm(r["cliente_nome"]) or s in _norm(r["cpf"]) or s in _norm(r["numero"]) or s in _norm(r["iccid"]) or s in _norm(r["email"] or "")]
    if status:
        result = [r for r in result if r["status_linha"] == status]
    if canal:
        result = [r for r in result if _norm(r["canal"]) == _norm(canal)]
    if bloqueio:
        result = [r for r in result if _norm(r["status_chip"]) == _norm(bloqueio)]

    # Ordenar por nome
    result.sort(key=lambda r: _norm(r["cliente_nome"]))

    # Resumo - considera flags incluir_custo e incluir_lucro
    total_receita = sum(r["valor_liquido"] for r in result if r.get("incluir_lucro"))
    total_custo = sum(r["custo"] for r in result if r.get("incluir_custo"))
    total_lucro = total_receita - total_custo
    margem_pct = round((total_lucro / total_receita * 100), 2) if total_receita > 0 else 0
    ativas = sum(1 for r in result if r["status_linha"] == "ativo")
    suspensas = sum(1 for r in result if r["status_linha"] == "suspenso")
    canceladas = sum(1 for r in result if r["status_linha"] == "cancelado")

    # Custos Fixos do Painel (VPS, dominio, Asaas, etc) - somente ativos
    custos_fixos_docs = await db.custos_fixos.find({"ativo": True}).to_list(100)
    custo_fixo_total = sum(float(d.get("valor") or 0) for d in custos_fixos_docs)
    custo_total_geral = total_custo + custo_fixo_total

    resumo = {
        "total_linhas": len(result),
        "ativas": ativas,
        "suspensas": suspensas,
        "canceladas": canceladas,
        "receita": round(total_receita, 2),
        "custo": round(total_custo, 2),
        "custo_fixo": round(custo_fixo_total, 2),
        "custo_total": round(custo_total_geral, 2),
        "lucro": round(total_lucro, 2),
        "margem_pct": margem_pct,
    }
    return {"linhas": result, "resumo": resumo}


def _resumo_vazio():
    return {"total_linhas": 0, "ativas": 0, "suspensas": 0, "canceladas": 0, "receita": 0, "custo": 0, "lucro": 0, "margem_pct": 0}


@router.patch("/linha/{linha_id}")
async def atualizar_linha_inline(linha_id: str, data: LinhaOperacionalUpdate, request: Request):
    user = await _ctx["require_admin"](request)
    db = _ctx["db"]
    if not ObjectId.is_valid(linha_id):
        raise HTTPException(status_code=400, detail="ID invalido")
    linha = await db.linhas.find_one({"_id": ObjectId(linha_id)})
    if not linha:
        raise HTTPException(status_code=404, detail="Linha nao encontrada")

    update_linha = {}
    if data.observacoes is not None:
        update_linha["observacoes"] = data.observacoes
    if data.proxima_recarga is not None:
        update_linha["proxima_recarga"] = data.proxima_recarga
    if data.status_chip is not None:
        update_linha["status_chip"] = data.status_chip
    if data.complemento is not None:
        update_linha["complemento"] = data.complemento
    if data.incluir_custo is not None:
        update_linha["incluir_custo"] = data.incluir_custo
    if data.incluir_lucro is not None:
        update_linha["incluir_lucro"] = data.incluir_lucro
    if data.desconto is not None:
        val = float(data.desconto)
        update_linha["desconto"] = val if val >= 0 else 0.0
    if update_linha:
        await db.linhas.update_one({"_id": ObjectId(linha_id)}, {"$set": update_linha})

    if data.canal is not None and linha.get("cliente_id") and ObjectId.is_valid(linha["cliente_id"]):
        await db.clientes.update_one({"_id": ObjectId(linha["cliente_id"])}, {"$set": {"canal": data.canal}})

    await _ctx["create_log"]("operacional", f"Linha {linha_id} atualizada inline", user["id"], user["name"])
    return {"success": True, "updated": {**update_linha, **({"canal": data.canal} if data.canal is not None else {})}}


@router.get("/export")
async def exportar_excel(request: Request):
    """Exporta a planilha consolidada em XLSX."""
    await _ctx["require_admin"](request)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    # Reusa a planilha
    data = await planilha_consolidada(request)
    linhas = data["linhas"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilha Operacional"

    headers = [
        "Cliente", "CPF", "Telefone", "Email", "Canal",
        "ICCID", "Numero", "Status Linha", "Status Chip",
        "Recarga Ta", "Prox. Boleto",
        "Oferta", "Plano", "Franquia",
        "Valor (R$)", "Desconto (R$)", "Valor Liq. (R$)", "Custo (R$)", "Lucro (R$)", "Margem %",
        "Categoria",
        "Ultima Cobranca Venc", "Ultima Cobranca Status", "Tipo Boleto",
        "Cobrancas Total", "Pagas", "Pendentes",
        "Endereco", "Observacoes Cliente", "Observacoes Linha",
    ]
    ws.append(headers)
    # Style header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in linhas:
        ws.append([
            r["cliente_nome"], r["cpf"], r["telefone_cliente"], r.get("email") or "", r["canal"],
            r["iccid"], r["numero"], r["status_linha"], r["status_chip"],
            r.get("expirar_dados") or "", r.get("proxima_recarga") or "",
            r["oferta_nome"], r["plano_nome"], r["franquia"],
            r["valor"], r.get("desconto", 0), r.get("valor_liquido", r["valor"]), r["custo"], r["lucro"], r["margem_pct"],
            r["categoria"],
            r.get("ultima_cobranca_venc") or "", r.get("ultima_cobranca_status") or "", r.get("ultima_cobranca_tipo") or "",
            r["cobrancas_total"], r["cobrancas_pagas"], r["cobrancas_pendentes"],
            r["endereco"], r["observacoes_cliente"], r["observacoes_linha"],
        ])

    # Auto width
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=10)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 40)

    # Summary sheet
    ws2 = wb.create_sheet("Resumo")
    resumo = data["resumo"]
    ws2.append(["Indicador", "Valor"])
    ws2.append(["Total Linhas", resumo["total_linhas"]])
    ws2.append(["Ativas", resumo["ativas"]])
    ws2.append(["Suspensas", resumo["suspensas"]])
    ws2.append(["Canceladas", resumo["canceladas"]])
    ws2.append(["Receita (R$)", resumo["receita"]])
    ws2.append(["Custos - Variavel (R$)", resumo["custo"]])
    ws2.append(["Custos - Fixos Painel (R$)", resumo.get("custo_fixo", 0)])
    ws2.append(["Custo Total (R$)", resumo.get("custo_total", resumo["custo"])])
    ws2.append(["Lucro (R$)", resumo["lucro"]])
    ws2.append(["Margem %", resumo["margem_pct"]])
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"planilha-operacional-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.post("/importar-excel")
async def importar_excel(request: Request, file: UploadFile = File(...)):
    """Importa clientes + observacoes de uma planilha Excel. Nao sobrescreve dados existentes nao-nulos (merge)."""
    user = await _ctx["require_admin"](request)
    db = _ctx["db"]
    import openpyxl
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Arquivo invalido: {e}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Planilha vazia")

    # Heuristica: detecta colunas pelo primeiro row que contenha texto que case nome/cpf
    header_idx = 0
    for i, row in enumerate(rows[:5]):
        joined = " ".join(str(c or "").lower() for c in row)
        if any(k in joined for k in ["assinante", "cliente", "nome"]):
            header_idx = i
            break

    headers = [(_norm(str(c or ""))) for c in rows[header_idx]]

    def col(row, *names):
        for name in names:
            for i, h in enumerate(headers):
                if name in h:
                    return row[i] if i < len(row) else None
        return None

    imported = 0
    updated_count = 0
    errors = []

    for row in rows[header_idx + 1:]:
        if not any(row):
            continue
        nome = str(col(row, "assinante", "cliente", "nome") or "").strip()
        cpf_raw = col(row, "cpf", "documento")
        cpf = re.sub(r"\D", "", str(cpf_raw or ""))
        tel = re.sub(r"\D", "", str(col(row, "chip", "telefone", "numero") or ""))
        endereco = str(col(row, "endereco") or "").strip() or None
        obs = str(col(row, "observa", "status") or "").strip() or None
        valor_raw = col(row, "valor")
        plano_txt = str(col(row, "plano") or "").strip() or None

        if not nome:
            continue

        # Match por CPF (preferido), senao por telefone, senao por nome
        existing = None
        if cpf and len(cpf) >= 11:
            existing = await db.clientes.find_one({"documento": cpf})
        if not existing and tel and len(tel) >= 10:
            existing = await db.clientes.find_one({"telefone": {"$regex": tel[-9:]}})
        if not existing:
            existing = await db.clientes.find_one({"nome": {"$regex": f"^{re.escape(nome)}$", "$options": "i"}})

        if existing:
            upd = {}
            if obs and not existing.get("observacoes"):
                upd["observacoes"] = obs
            if endereco and not existing.get("endereco"):
                upd["endereco"] = endereco
            if upd:
                await db.clientes.update_one({"_id": existing["_id"]}, {"$set": upd})
                updated_count += 1
        else:
            if not cpf or len(cpf) < 11:
                errors.append(f"{nome}: CPF invalido, pulado")
                continue
            try:
                doc = {
                    "nome": nome, "documento": cpf, "tipo_pessoa": "pf",
                    "telefone": tel or "",
                    "endereco": endereco,
                    "observacoes": obs,
                    "status": "ativo",
                    "created_at": datetime.now(timezone.utc),
                    "imported_from_excel": True,
                }
                await db.clientes.insert_one(doc)
                imported += 1
            except Exception as e:
                errors.append(f"{nome}: {e}")

    await _ctx["create_log"]("operacional", f"Import Excel: {imported} novos, {updated_count} atualizados", user["id"], user["name"])
    return {"imported": imported, "updated": updated_count, "errors": errors[:50], "total_errors": len(errors)}


@router.post("/atualizar-expirar-dados/{iccid}")
async def atualizar_expirar_dados(iccid: str, request: Request):
    """Consulta Ta Telecom e cacheia a data de expiracao dos dados em linhas.expirar_dados."""
    user = await _ctx["require_admin"](request)
    db = _ctx["db"]
    # Import local para nao criar import circular
    from services.operadora_service import operadora_service
    try:
        resp = await operadora_service.consultar_linha(iccid, db=db, user_id=user["id"], user_name=user["name"])
        if not resp or not resp.success:
            raise HTTPException(status_code=502, detail=f"Falha Ta Telecom: {getattr(resp, 'message', '') or getattr(resp, 'status', '?')}")
        data = resp.data if isinstance(resp.data, dict) else {}
        inner = data.get("data") if isinstance(data.get("data"), dict) else data
        data_at = _parse_data_br(inner.get("data_ativacao") or data.get("data_ativacao"))
        data_bl_raw = _parse_data_br(inner.get("data_bloqueio") or data.get("data_bloqueio"))
        expirar_direto = _parse_data_br(
            inner.get("data_expiracao") or inner.get("dataExpiracao")
            or inner.get("expira_em") or inner.get("validity")
            or inner.get("data_validade") or inner.get("validade")
        )
        # Usa resolver que garante data FUTURA (descarta datas passadas)
        expirar = _resolver_proxima_recarga(
            data_ativacao=data_at, data_bloqueio=data_bl_raw,
            data_expiracao_direta=expirar_direto,
        )
        chip = await db.chips.find_one({"iccid": iccid})
        if chip:
            await db.linhas.update_many({"chip_id": str(chip["_id"])}, {"$set": {"expirar_dados": expirar, "expirar_dados_updated_at": datetime.now(timezone.utc)}})
        await _ctx["create_log"]("operacional", f"Expirar dados atualizado via TaTelecom: ICCID {iccid}", user["id"], user["name"])
        return {"iccid": iccid, "expirar_dados": expirar, "raw": inner}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== CUSTOS ====================
@router.get("/ofertas-com-stats")
async def ofertas_com_stats(request: Request):
    """Lista todas ofertas com numero de linhas ativas e soma de receita/custo/lucro para gestao de custos."""
    await _ctx["require_admin"](request)
    db = _ctx["db"]
    ofertas = await db.ofertas.find({}).to_list(1000)
    # Planos
    plano_ids = list({o.get("plano_id") for o in ofertas if o.get("plano_id")})
    planos_map = {}
    if plano_ids:
        pls = await db.planos.find({"_id": {"$in": [ObjectId(p) for p in plano_ids if ObjectId.is_valid(p)]}}).to_list(500)
        planos_map = {str(p["_id"]): p for p in pls}

    # Contar linhas ativas por oferta (ou por plano se linha nao tem oferta_id)
    linhas = await db.linhas.find({"status": {"$in": ["ativo", "suspenso"]}}).to_list(5000)
    count_por_oferta = {}
    count_por_plano = {}
    for l in linhas:
        oid = l.get("oferta_id")
        if oid:
            count_por_oferta[oid] = count_por_oferta.get(oid, 0) + 1
        else:
            pid = l.get("plano_id")
            if pid:
                count_por_plano[pid] = count_por_plano.get(pid, 0) + 1

    result = []
    for o in ofertas:
        oid = str(o["_id"])
        plano = planos_map.get(o.get("plano_id") or "", {})
        direct = count_por_oferta.get(oid, 0)
        # Se for a oferta default de um plano, contar linhas do plano sem oferta_id
        indirect = count_por_plano.get(o.get("plano_id"), 0) if direct == 0 else 0
        total_linhas = direct + indirect
        valor = o.get("valor", 0) or 0
        custo = o.get("custo", 0) or 0
        result.append({
            "id": oid,
            "nome": o["nome"],
            "categoria": o.get("categoria", "movel"),
            "plano_nome": plano.get("nome", ""),
            "franquia": plano.get("franquia", ""),
            "valor": valor,
            "custo": custo,
            "lucro": valor - custo,
            "margem_pct": round((valor - custo) / valor * 100, 2) if valor > 0 else 0,
            "linhas_ativas": total_linhas,
            "receita_total": round(valor * total_linhas, 2),
            "custo_total": round(custo * total_linhas, 2),
            "lucro_total": round((valor - custo) * total_linhas, 2),
            "ativo": o.get("ativo", True),
        })
    result.sort(key=lambda x: _sort_key_plano(x["plano_nome"] or x["nome"] or "", x.get("franquia") or ""))
    return result


@router.patch("/oferta/{oferta_id}/custo")
async def atualizar_custo_oferta(oferta_id: str, data: OfertaCustoUpdate, request: Request):
    user = await _ctx["require_admin"](request)
    db = _ctx["db"]
    if not ObjectId.is_valid(oferta_id):
        raise HTTPException(status_code=400, detail="ID invalido")
    oferta = await db.ofertas.find_one({"_id": ObjectId(oferta_id)})
    if not oferta:
        raise HTTPException(status_code=404, detail="Oferta nao encontrada")
    if data.custo < 0:
        raise HTTPException(status_code=400, detail="Custo nao pode ser negativo")
    await db.ofertas.update_one({"_id": ObjectId(oferta_id)}, {"$set": {"custo": data.custo}})
    await _ctx["create_log"]("operacional", f"Custo da oferta '{oferta['nome']}' atualizado: R$ {data.custo:.2f}", user["id"], user["name"])
    return {"success": True, "oferta_id": oferta_id, "custo": data.custo}


@router.post("/custos/batch")
async def atualizar_custos_batch(data: CustosBatchUpdate, request: Request):
    """Atualiza custos de varias ofertas de uma vez."""
    user = await _ctx["require_admin"](request)
    db = _ctx["db"]
    updated = 0
    errors = []
    for oferta_id, custo in data.custos.items():
        try:
            if not ObjectId.is_valid(oferta_id):
                errors.append(f"{oferta_id}: ID invalido")
                continue
            custo_f = float(custo)
            if custo_f < 0:
                errors.append(f"{oferta_id}: custo negativo")
                continue
            r = await db.ofertas.update_one({"_id": ObjectId(oferta_id)}, {"$set": {"custo": custo_f}})
            if r.modified_count:
                updated += 1
        except Exception as e:
            errors.append(f"{oferta_id}: {e}")
    await _ctx["create_log"]("operacional", f"Custos em lote: {updated} ofertas atualizadas", user["id"], user["name"])
    return {"updated": updated, "errors": errors}


# ==================== SINCRONIZACAO AUTOMATICA TA TELECOM ====================
async def _auto_sync_linhas_stale(linha_ids: list):
    """Sincroniza com Ta Telecom apenas as linhas que precisam (data passada/stale).

    Executa em background, disparado pelo GET /planilha. NAO bloqueia a resposta.
    Cada consulta tem throttle de 0.4s para nao estourar rate limit.
    """
    db = _ctx["db"]
    from services.operadora_service import operadora_service
    import asyncio as _asyncio

    try:
        obj_ids = [ObjectId(x) for x in linha_ids if ObjectId.is_valid(x)]
        if not obj_ids:
            return
        linhas = await db.linhas.find({"_id": {"$in": obj_ids}}).to_list(len(obj_ids))
        chip_ids = list({l["chip_id"] for l in linhas if l.get("chip_id") and ObjectId.is_valid(l["chip_id"])})
        chips_map = {}
        if chip_ids:
            chs = await db.chips.find({"_id": {"$in": [ObjectId(c) for c in chip_ids]}}).to_list(len(chip_ids))
            chips_map = {str(c["_id"]): c for c in chs}

        for l in linhas:
            chip = chips_map.get(l.get("chip_id") or "")
            if not chip or not chip.get("iccid"):
                continue
            iccid = chip["iccid"]
            try:
                await _asyncio.sleep(0.4)  # throttle rate limit
                resp = await operadora_service.consultar_linha(iccid, db=db, user_id="auto-sync", user_name="auto-sync")
                if not resp or not resp.success:
                    continue
                data = resp.data if isinstance(resp.data, dict) else {}
                inner = data.get("data") if isinstance(data.get("data"), dict) else data
                data_at = _parse_data_br(inner.get("data_ativacao") or data.get("data_ativacao"))
                data_bl_raw = _parse_data_br(inner.get("data_bloqueio") or data.get("data_bloqueio"))
                expirar_direto = _parse_data_br(
                    inner.get("data_expiracao") or inner.get("dataExpiracao")
                    or inner.get("validade") or inner.get("data_validade")
                )
                expirar = _resolver_proxima_recarga(
                    data_ativacao=data_at, data_bloqueio=data_bl_raw,
                    data_expiracao_direta=expirar_direto,
                )
                status_raw = str(
                    inner.get("status") or inner.get("situacao") or data.get("status") or ""
                ).upper().strip()
                status_map = {
                    "FUNCIONAL": "Ativo", "ATIVO": "Ativo", "FS": "Ativo",
                    "NOVO": "Novo", "NP": "Novo",
                    "BLOQUEADO_PARCIAL": "Bloq. Parcial",
                    "BLOQUEADO_TOTAL": "Bloqueado", "BLOQUEADO": "Bloqueado",
                    "CANCELADO": "Cancelado", "SUSPENSO": "Suspenso",
                    "PENDENTE": "Pendente", "PENDING": "Pendente",
                }
                status_chip_sigla = status_map.get(status_raw, status_raw.title() if status_raw else None)
                update = {"expirar_dados_updated_at": datetime.now(timezone.utc)}
                if expirar:
                    update["expirar_dados"] = expirar
                if status_chip_sigla:
                    update["status_chip"] = status_chip_sigla
                if len(update) > 1:
                    await db.linhas.update_one({"_id": l["_id"]}, {"$set": update})
            except Exception as e:
                logger.warning(f"Auto-sync falhou para ICCID {iccid}: {e}")
                continue
    except Exception as e:
        logger.error(f"Erro no auto-sync de linhas: {e}")


async def _run_sync_tatelecom_bg(user_id: str, user_name: str):
    """Executa sync com Ta Telecom em background e grava progresso em db.sync_jobs."""
    db = _ctx["db"]
    from services.operadora_service import operadora_service
    import asyncio as _asyncio

    job_id = f"tatelecom-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    await db.sync_jobs.insert_one({
        "job_id": job_id,
        "tipo": "tatelecom",
        "status": "running",
        "iniciado_em": datetime.now(timezone.utc),
        "total": 0, "atualizadas": 0, "erros": 0,
        "user_id": user_id,
    })

    try:
        linhas = await db.linhas.find({"status": {"$in": ["ativo", "suspenso"]}}).to_list(5000)
        chip_ids = list({l["chip_id"] for l in linhas if l.get("chip_id") and ObjectId.is_valid(l["chip_id"])})
        chips_map = {}
        if chip_ids:
            chs = await db.chips.find({"_id": {"$in": [ObjectId(c) for c in chip_ids]}}).to_list(5000)
            chips_map = {str(c["_id"]): c for c in chs}

        await db.sync_jobs.update_one({"job_id": job_id}, {"$set": {"total": len(linhas)}})

        atualizados = 0
        erros_count = 0
        erros_lista = []
        nao_encontrados = 0

        for idx, l in enumerate(linhas):
            chip = chips_map.get(l.get("chip_id") or "")
            if not chip or not chip.get("iccid"):
                nao_encontrados += 1
                continue
            iccid = chip["iccid"]
            if idx > 0 and idx % 10 == 0:
                await _asyncio.sleep(2)
            else:
                await _asyncio.sleep(0.3)
            try:
                resp = await operadora_service.consultar_linha(iccid, db=db, user_id=user_id, user_name=user_name)
                if resp and not resp.success and "429" in str(getattr(resp, "message", "")):
                    await _asyncio.sleep(5)
                    resp = await operadora_service.consultar_linha(iccid, db=db, user_id=user_id, user_name=user_name)
                if not resp or not resp.success:
                    erros_count += 1
                    if len(erros_lista) < 30:
                        erros_lista.append(f"{iccid}: {getattr(resp, 'message', '') or getattr(resp, 'status', '?')}")
                    continue
                data = resp.data if isinstance(resp.data, dict) else {}
                inner = data.get("data") if isinstance(data.get("data"), dict) else data
                # Ta Telecom retorna data_ativacao e data_bloqueio. Quando o chip nunca
                # foi bloqueado, data_bloqueio == data_ativacao. Para "Proxima Recarga"
                # correta, usamos _calc_proxima_recarga que trata esse caso.
                data_at = _parse_data_br(inner.get("data_ativacao") or data.get("data_ativacao"))
                data_bl_raw = _parse_data_br(inner.get("data_bloqueio") or data.get("data_bloqueio"))
                expirar_direto = _parse_data_br(
                    inner.get("data_expiracao") or inner.get("dataExpiracao")
                    or inner.get("validade") or inner.get("data_validade")
                )
                # Resolver que SEMPRE retorna data futura (ignora datas passadas)
                expirar = _resolver_proxima_recarga(
                    data_ativacao=data_at, data_bloqueio=data_bl_raw,
                    data_expiracao_direta=expirar_direto,
                )
                status_raw = str(
                    inner.get("status") or inner.get("situacao") or data.get("status") or ""
                ).upper().strip()
                status_map = {
                    "FUNCIONAL": "Ativo", "ATIVO": "Ativo", "FS": "Ativo",
                    "NOVO": "Novo", "NP": "Novo",
                    "BLOQUEADO_PARCIAL": "Bloq. Parcial",
                    "BLOQUEADO_TOTAL": "Bloqueado", "BLOQUEADO": "Bloqueado",
                    "CANCELADO": "Cancelado", "SUSPENSO": "Suspenso",
                    "PENDENTE": "Pendente", "PENDING": "Pendente",
                }
                status_chip_sigla = status_map.get(status_raw, status_raw.title() if status_raw else None)
                update = {"expirar_dados_updated_at": datetime.now(timezone.utc)}
                if expirar:
                    update["expirar_dados"] = expirar
                if status_chip_sigla:
                    update["status_chip"] = status_chip_sigla
                if len(update) > 1:
                    await db.linhas.update_one({"_id": l["_id"]}, {"$set": update})
                    atualizados += 1
            except Exception as e:
                erros_count += 1
                if len(erros_lista) < 30:
                    erros_lista.append(f"{iccid}: {str(e)[:100]}")

            # Atualizar progresso a cada 5 linhas
            if (idx + 1) % 5 == 0 or idx == len(linhas) - 1:
                await db.sync_jobs.update_one({"job_id": job_id}, {"$set": {
                    "processadas": idx + 1, "atualizadas": atualizados, "erros": erros_count,
                }})

        await db.sync_jobs.update_one({"job_id": job_id}, {"$set": {
            "status": "completed",
            "finalizado_em": datetime.now(timezone.utc),
            "atualizadas": atualizados,
            "sem_chip": nao_encontrados,
            "erros": erros_count,
            "erros_lista": erros_lista,
        }})
        await _ctx["create_log"]("operacional", f"Sync TaTelecom batch: {atualizados}/{len(linhas)} atualizadas", user_id, user_name)
    except Exception as e:
        await db.sync_jobs.update_one({"job_id": job_id}, {"$set": {
            "status": "error",
            "finalizado_em": datetime.now(timezone.utc),
            "error_message": str(e)[:500],
        }})


@router.post("/sincronizar-tatelecom")
async def sincronizar_tatelecom_batch(request: Request):
    """Dispara sync em background. Retorna imediatamente. Use GET /sync-status para acompanhar."""
    user = await _ctx["require_admin"](request)
    from services.operadora_service import operadora_service
    if getattr(operadora_service, "use_mock", True):
        raise HTTPException(status_code=400, detail="Ta Telecom nao configurado (em modo mock)")

    # Checar se ja tem um job rodando
    db = _ctx["db"]
    existing = await db.sync_jobs.find_one({"tipo": "tatelecom", "status": "running"})
    if existing:
        return {
            "status": "already_running",
            "message": "Ja existe uma sincronizacao em andamento",
            "job_id": existing["job_id"],
        }

    import asyncio as _asyncio
    _asyncio.create_task(_run_sync_tatelecom_bg(user["id"], user["name"]))
    return {
        "status": "started",
        "message": "Sincronizacao iniciada em background. Use GET /sync-status/tatelecom para acompanhar.",
    }


@router.get("/sync-status/tatelecom")
async def sync_status_tatelecom(request: Request):
    """Retorna status do ultimo job de sync da Ta Telecom."""
    await _ctx["require_admin"](request)
    db = _ctx["db"]
    job = await db.sync_jobs.find_one({"tipo": "tatelecom"}, sort=[("iniciado_em", -1)])
    if not job:
        return {"status": "never_run"}
    return {
        "job_id": job.get("job_id"),
        "status": job.get("status"),
        "total": job.get("total", 0),
        "processadas": job.get("processadas", 0),
        "atualizadas": job.get("atualizadas", 0),
        "erros": job.get("erros", 0),
        "sem_chip": job.get("sem_chip", 0),
        "iniciado_em": job.get("iniciado_em").isoformat() if job.get("iniciado_em") else None,
        "finalizado_em": job.get("finalizado_em").isoformat() if job.get("finalizado_em") else None,
        "error_message": job.get("error_message"),
        "erros_lista": job.get("erros_lista", [])[:10],
    }


# ==================== AUTO-PREENCHER CANAL ====================
@router.post("/auto-canal")
async def auto_preencher_canal(request: Request):
    """Preenche canal automaticamente:
       - Cliente veio de self-service sem revendedor -> 'Proprio'
       - Cliente vinculado a chip de revendedor -> 'Revendedor'
    Nao sobrescreve canal ja definido manualmente.
    """
    user = await _ctx["require_admin"](request)
    db = _ctx["db"]

    # 1. Clientes com chip de revendedor
    chips_rev = await db.chips.find({"revendedor_id": {"$exists": True, "$ne": None}}, {"cliente_id": 1}).to_list(5000)
    cliente_ids_rev = list({c.get("cliente_id") for c in chips_rev if c.get("cliente_id")})
    updated_rev = 0
    if cliente_ids_rev:
        r = await db.clientes.update_many(
            {"_id": {"$in": [ObjectId(c) for c in cliente_ids_rev if ObjectId.is_valid(c)]},
             "$or": [{"canal": None}, {"canal": ""}, {"canal": {"$exists": False}}]},
            {"$set": {"canal": "Revendedor"}}
        )
        updated_rev = r.modified_count

    # 2. Clientes de self-service (sem revendedor) -> Proprio
    ss_ativacoes = await db.ativacoes_selfservice.find({"cliente_id": {"$exists": True, "$ne": None}}, {"cliente_id": 1}).to_list(5000)
    cliente_ids_ss = list({s.get("cliente_id") for s in ss_ativacoes if s.get("cliente_id")})
    updated_ss = 0
    if cliente_ids_ss:
        r = await db.clientes.update_many(
            {"_id": {"$in": [ObjectId(c) for c in cliente_ids_ss if ObjectId.is_valid(c)]},
             "$or": [{"canal": None}, {"canal": ""}, {"canal": {"$exists": False}}]},
            {"$set": {"canal": "Proprio"}}
        )
        updated_ss = r.modified_count

    await _ctx["create_log"]("operacional", f"Auto canal: {updated_rev} Revendedor, {updated_ss} Proprio", user["id"], user["name"])
    return {"updated_revendedor": updated_rev, "updated_proprio": updated_ss}


# ==================== AUTO PROXIMA RECARGA ====================
@router.post("/auto-proxima-recarga")
async def auto_proxima_recarga(request: Request):
    """Calcula proxima_recarga a partir do ultimo boleto pago do cliente:
       vencimento do ultimo pago + 30 dias. Nao sobrescreve valor ja preenchido manualmente.
    """
    user = await _ctx["require_admin"](request)
    db = _ctx["db"]
    from datetime import timedelta

    linhas = await db.linhas.find({"status": {"$in": ["ativo", "suspenso"]}}).to_list(5000)
    # Ultima cobranca paga por cliente
    pagas = await db.cobrancas.find({"status": {"$in": ["RECEIVED", "CONFIRMED", "RECEIVED_IN_CASH"]}}, {"cliente_id": 1, "vencimento": 1}).to_list(20000)
    ultima_paga_por_cliente = {}
    for c in pagas:
        cid = c.get("cliente_id")
        venc = c.get("vencimento")
        if not cid or not venc:
            continue
        if cid not in ultima_paga_por_cliente or venc > ultima_paga_por_cliente[cid]:
            ultima_paga_por_cliente[cid] = venc

    updated = 0
    for l in linhas:
        if l.get("proxima_recarga"):
            continue  # nao sobrescreve manual
        cid = l.get("cliente_id")
        last_venc = ultima_paga_por_cliente.get(cid)
        if not last_venc:
            continue
        try:
            dt = datetime.fromisoformat(str(last_venc)[:10])
            prox = dt + timedelta(days=30)
            await db.linhas.update_one({"_id": l["_id"]}, {"$set": {"proxima_recarga": prox.strftime("%Y-%m-%d")}})
            updated += 1
        except Exception:
            continue

    await _ctx["create_log"]("operacional", f"Auto proxima recarga: {updated} linhas atualizadas", user["id"], user["name"])
    return {"updated": updated, "total_linhas": len(linhas)}


# ==================== LIMPEZA DE DATAS PASSADAS ====================
@router.post("/limpar-datas-passadas")
async def limpar_datas_passadas(request: Request):
    """Varre linhas com `expirar_dados` no passado e recalcula via ciclo de 30 dias.

    Utiliza `data_ativacao` do chip (se disponivel) OU a ultima cobranca paga.
    Chamado automaticamente ao abrir a planilha operacional e pelo scheduler diario.
    """
    user = await _ctx["require_admin"](request)
    db = _ctx["db"]
    from datetime import timedelta, date as _date
    hoje = _date.today()
    hoje_iso = hoje.strftime("%Y-%m-%d")

    # Pega linhas com expirar_dados no passado
    linhas_passadas = await db.linhas.find({
        "expirar_dados": {"$ne": None, "$lt": hoje_iso},
    }).to_list(10000)

    if not linhas_passadas:
        return {"corrigidas": 0, "mensagem": "Nenhuma linha com data passada"}

    # Pega data_ativacao dos chips envolvidos
    chip_ids = list({l["chip_id"] for l in linhas_passadas if l.get("chip_id") and ObjectId.is_valid(l["chip_id"])})
    chips_map = {}
    if chip_ids:
        chs = await db.chips.find({"_id": {"$in": [ObjectId(c) for c in chip_ids]}}, {"data_ativacao": 1}).to_list(10000)
        chips_map = {str(c["_id"]): c for c in chs}

    # Ultima cobranca paga por cliente (fallback)
    pagas = await db.cobrancas.find(
        {"status": {"$in": ["RECEIVED", "CONFIRMED", "RECEIVED_IN_CASH"]}},
        {"cliente_id": 1, "vencimento": 1, "pago_em": 1},
    ).to_list(30000)
    ultima_paga_por_cliente = {}
    for c in pagas:
        cid = c.get("cliente_id")
        venc = c.get("vencimento") or c.get("pago_em")
        if not cid or not venc:
            continue
        venc_str = str(venc)[:10]
        if cid not in ultima_paga_por_cliente or venc_str > ultima_paga_por_cliente[cid]:
            ultima_paga_por_cliente[cid] = venc_str

    corrigidas = 0
    for l in linhas_passadas:
        expirar_atual = l.get("expirar_dados")
        if not expirar_atual:
            continue
        # 1. Tenta avancar 30 dias a partir da data passada ate achar uma futura
        try:
            dt = datetime.fromisoformat(expirar_atual[:10]).date()
            while dt <= hoje:
                dt = dt + timedelta(days=30)
            nova_data = dt.strftime("%Y-%m-%d")
        except Exception:
            nova_data = None

        # 2. Se ainda nao tem data, tenta via ultima cobranca paga
        if not nova_data:
            cid = l.get("cliente_id")
            last_venc = ultima_paga_por_cliente.get(cid)
            if last_venc:
                try:
                    dt = datetime.fromisoformat(last_venc).date() + timedelta(days=30)
                    while dt <= hoje:
                        dt = dt + timedelta(days=30)
                    nova_data = dt.strftime("%Y-%m-%d")
                except Exception:
                    pass

        if nova_data and nova_data != expirar_atual:
            await db.linhas.update_one(
                {"_id": l["_id"]},
                {"$set": {"expirar_dados": nova_data, "expirar_dados_updated_at": datetime.now(timezone.utc)}},
            )
            corrigidas += 1

    await _ctx["create_log"]("operacional", f"Limpeza de datas passadas: {corrigidas}/{len(linhas_passadas)} linhas corrigidas", user["id"], user["name"])
    return {"corrigidas": corrigidas, "total_analisadas": len(linhas_passadas)}


# ==================== CUSTO POR PLANO (aplica em todas ofertas) ====================
@router.patch("/plano/{plano_id}/custo")
async def atualizar_custo_plano(plano_id: str, data: PlanoCustoUpdate, request: Request):
    """Aplica o mesmo custo a TODAS as ofertas do plano informado."""
    user = await _ctx["require_admin"](request)
    db = _ctx["db"]
    if not ObjectId.is_valid(plano_id):
        raise HTTPException(status_code=400, detail="ID invalido")
    plano = await db.planos.find_one({"_id": ObjectId(plano_id)})
    if not plano:
        raise HTTPException(status_code=404, detail="Plano nao encontrado")
    if data.custo < 0:
        raise HTTPException(status_code=400, detail="Custo nao pode ser negativo")
    r = await db.ofertas.update_many({"plano_id": plano_id}, {"$set": {"custo": data.custo}})
    await _ctx["create_log"]("operacional", f"Custo do plano '{plano['nome']}' aplicado a {r.modified_count} ofertas: R$ {data.custo:.2f}", user["id"], user["name"])
    return {"success": True, "plano_id": plano_id, "ofertas_atualizadas": r.modified_count, "custo": data.custo}


@router.get("/planos-com-stats")
async def planos_com_stats(request: Request):
    """Lista planos com resumo de ofertas e custo base."""
    await _ctx["require_admin"](request)
    db = _ctx["db"]
    planos = await db.planos.find({}).to_list(500)
    result = []
    for p in planos:
        pid = str(p["_id"])
        ofertas = await db.ofertas.find({"plano_id": pid}).to_list(100)
        linhas = await db.linhas.count_documents({"plano_id": pid, "status": {"$in": ["ativo", "suspenso"]}})
        # Se ofertas tem custos diferentes, usa o do primeiro; se iguais, usa
        custos = list({o.get("custo", 0) for o in ofertas})
        custo_base = custos[0] if len(custos) == 1 else (ofertas[0].get("custo", 0) if ofertas else 0)
        result.append({
            "id": pid,
            "nome": p["nome"],
            "franquia": p.get("franquia", ""),
            "ofertas_count": len(ofertas),
            "linhas_ativas": linhas,
            "custo_base": custo_base,
            "custos_diferentes": len(custos) > 1,
        })
    result.sort(key=lambda x: _sort_key_plano(x["nome"] or "", x.get("franquia") or ""))
    return result


# ==================== CUSTOS FIXOS (painel, VPS, etc) ====================
@router.get("/custos-fixos")
async def listar_custos_fixos(request: Request):
    await _ctx["require_admin"](request)
    db = _ctx["db"]
    docs = await db.custos_fixos.find({}).sort("nome", 1).to_list(100)
    return [{"id": str(d["_id"]), "nome": d["nome"], "valor": d.get("valor", 0), "ativo": d.get("ativo", True)} for d in docs]


@router.post("/custos-fixos")
async def criar_custo_fixo(data: CustoFixoCreate, request: Request):
    user = await _ctx["require_admin"](request)
    db = _ctx["db"]
    if data.valor < 0:
        raise HTTPException(status_code=400, detail="Valor nao pode ser negativo")
    doc = {"nome": data.nome, "valor": data.valor, "ativo": data.ativo, "created_at": datetime.now(timezone.utc)}
    r = await db.custos_fixos.insert_one(doc)
    await _ctx["create_log"]("operacional", f"Custo fixo criado: {data.nome} R$ {data.valor:.2f}", user["id"], user["name"])
    return {"id": str(r.inserted_id), "nome": data.nome, "valor": data.valor, "ativo": data.ativo}


@router.patch("/custos-fixos/{custo_id}")
async def atualizar_custo_fixo(custo_id: str, data: CustoFixoUpdate, request: Request):
    user = await _ctx["require_admin"](request)
    db = _ctx["db"]
    if not ObjectId.is_valid(custo_id):
        raise HTTPException(status_code=400, detail="ID invalido")
    update = {k: v for k, v in data.dict(exclude_none=True).items()}
    if "valor" in update and update["valor"] < 0:
        raise HTTPException(status_code=400, detail="Valor nao pode ser negativo")
    if not update:
        return {"success": True}
    await db.custos_fixos.update_one({"_id": ObjectId(custo_id)}, {"$set": update})
    await _ctx["create_log"]("operacional", f"Custo fixo {custo_id} atualizado", user["id"], user["name"])
    return {"success": True}


@router.delete("/custos-fixos/{custo_id}")
async def deletar_custo_fixo(custo_id: str, request: Request):
    user = await _ctx["require_admin"](request)
    db = _ctx["db"]
    if not ObjectId.is_valid(custo_id):
        raise HTTPException(status_code=400, detail="ID invalido")
    await db.custos_fixos.delete_one({"_id": ObjectId(custo_id)})
    await _ctx["create_log"]("operacional", f"Custo fixo {custo_id} removido", user["id"], user["name"])
    return {"success": True}


# ==================== RESUMO FINANCEIRO COMPLETO ====================
@router.get("/resumo-financeiro")
async def resumo_financeiro(request: Request):
    """Calcula receita total, custo variavel (ofertas*linhas), custo fixo (painel) e lucro final."""
    await _ctx["require_admin"](request)
    db = _ctx["db"]
    # Receita + custo variavel (reusa logica de ofertas-com-stats)
    ofertas = await db.ofertas.find({}).to_list(1000)
    # Pega TODAS as linhas ativas+suspensas+bloqueadas (pois podem ter flag custom)
    linhas = await db.linhas.find({}).to_list(5000)
    # Mapeamento oferta_id -> oferta para acesso rapido
    ofertas_map = {str(o["_id"]): o for o in ofertas}

    # Indice plano_id -> primeira oferta ativa (fallback quando linha nao tem oferta_id)
    oferta_por_plano = {}
    for o in ofertas:
        pid = o.get("plano_id")
        if pid and pid not in oferta_por_plano and o.get("ativo", True):
            oferta_por_plano[pid] = o

    receita = 0.0
    custo_variavel = 0.0
    for l in linhas:
        status = l.get("status", "")
        default_incluir = status == "ativo"
        incluir_custo = l.get("incluir_custo")
        incluir_lucro = l.get("incluir_lucro")
        if incluir_custo is None:
            incluir_custo = default_incluir
        if incluir_lucro is None:
            incluir_lucro = default_incluir

        oferta = ofertas_map.get(l.get("oferta_id") or "")
        if not oferta and l.get("plano_id"):
            oferta = oferta_por_plano.get(l["plano_id"])
        if not oferta:
            continue

        if incluir_lucro:
            receita += oferta.get("valor", 0) or 0
        if incluir_custo:
            custo_variavel += oferta.get("custo", 0) or 0

    # Custos fixos
    custos_fixos_docs = await db.custos_fixos.find({"ativo": True}).to_list(100)
    custo_fixo = sum(d.get("valor", 0) for d in custos_fixos_docs)

    custo_total = custo_variavel + custo_fixo
    lucro = receita - custo_total
    margem = (lucro / receita * 100) if receita > 0 else 0

    return {
        "receita": round(receita, 2),
        "custo_variavel": round(custo_variavel, 2),
        "custo_fixo": round(custo_fixo, 2),
        "custo_total": round(custo_total, 2),
        "lucro": round(lucro, 2),
        "margem_pct": round(margem, 2),
        "total_linhas": len(linhas),
        "total_custos_fixos": len(custos_fixos_docs),
    }

